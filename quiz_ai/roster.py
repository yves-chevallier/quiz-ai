"""
Utility helpers to load student rosters and match handwritten names against them.
"""

from __future__ import annotations

import csv
import re
import unicodedata
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

try:  # Optional dependency
    from openpyxl import load_workbook  # type: ignore
except ImportError:  # pragma: no cover - optional dependency
    load_workbook = None  # type: ignore


@dataclass
class StudentRecord:
    """Normalised representation of a student entry."""

    first_name: str
    last_name: str
    full_name: str
    email: Optional[str] = None
    raw: Dict[str, str] | None = None

    def display_name(self) -> str:
        if self.full_name:
            return self.full_name
        if self.first_name and self.last_name:
            return f"{self.first_name} {self.last_name}".strip()
        return self.first_name or self.last_name


@dataclass
class RosterCandidate:
    """Single candidate considered during matching."""

    name: str
    score: float
    confidence: str
    source: str


@dataclass
class RosterMatch:
    """Best roster match for a handwritten student name."""

    student: Optional[StudentRecord]
    score: float
    confidence: str
    source: str
    matched_variant: str
    candidates: Tuple[RosterCandidate, ...]


class RosterError(RuntimeError):
    """Raised when the roster cannot be parsed."""


_LAST_NAME_KEYS = {
    "nom",
    "lastname",
    "last_name",
    "familyname",
    "family_name",
    "surname",
}
_FIRST_NAME_KEYS = {
    "prenom",
    "prénom",
    "firstname",
    "first_name",
    "givenname",
    "given_name",
}
_FULL_NAME_KEYS = {"nom complet", "name", "full_name", "fullname"}
_EMAIL_KEYS = {"email", "e-mail", "mail"}


def load_roster(path: Path) -> List[StudentRecord]:
    """
    Load a roster file (CSV or XLSX) and return a list of student records.

    Raises
    ------
    RosterError
        If the file cannot be parsed or contains no usable student entries.
    """
    if not path.exists():
        raise RosterError(f"Roster file not found: {path}")

    suffix = path.suffix.lower()
    rows: List[Dict[str, str]]
    if suffix in {".csv", ".tsv"}:
        rows = _read_csv_rows(path)
    elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        rows = _read_xlsx_rows(path)
    else:
        raise RosterError(f"Unsupported roster format (expected CSV or XLSX): {path}")

    records: List[StudentRecord] = []
    for row in rows:
        record = _row_to_student(row)
        if record:
            records.append(record)

    if not records:
        raise RosterError(f"No student entries found in roster: {path}")
    return records


def match_student_name(
    handwritten_candidates: Sequence[Tuple[str, str]],
    roster: Sequence[StudentRecord],
    *,
    threshold: float = 0.65,
    max_candidates: int = 5,
) -> Optional[RosterMatch]:
    """
    Compare one or more handwritten name candidates against the roster.

    Parameters
    ----------
    handwritten_candidates:
        Sequence of tuples (value, source) where value is the raw string extracted
        from the scan and source describes where it came from (e.g., "student_name").
    roster:
        Student roster entries to match against.
    threshold:
        Minimum score for a match to be returned. Matches below this threshold
        are considered too uncertain and return ``None``.
    max_candidates:
        Number of alternative candidates to include in the returned structure.
    """
    inputs = list(_expand_handwritten_candidates(handwritten_candidates))
    if not inputs or not roster:
        return None

    roster_variants = [
        (record, list(_enumerate_record_variants(record))) for record in roster
    ]

    ranked: List[Tuple[float, StudentRecord, str, str, str]] = []
    best: Optional[Tuple[float, StudentRecord, str, str, str]] = None

    for normalized_value, tokens, original, source in inputs:
        for record, variants in roster_variants:
            for variant_norm, variant_tokens, variant_display in variants:
                score = _score_match(normalized_value, tokens, variant_norm, variant_tokens)
                ranked.append((score, record, variant_display, original, source))
                if best is None or score > best[0]:
                    best = (score, record, variant_display, original, source)

    if best is None or best[0] < threshold:
        return None

    score, record, variant_display, raw_candidate, source = best
    confidence = _score_to_confidence(score)
    candidates = _build_candidate_list(ranked, max_candidates)

    return RosterMatch(
        student=record,
        score=score,
        confidence=confidence,
        source=source,
        matched_variant=variant_display,
        candidates=candidates,
    )


def _build_candidate_list(
    ranked: Iterable[Tuple[float, StudentRecord, str, str, str]],
    limit: int,
) -> Tuple[RosterCandidate, ...]:
    seen: set[str] = set()
    ordered: List[RosterCandidate] = []
    for score, record, variant_display, _raw, source in sorted(ranked, key=lambda item: item[0], reverse=True):
        display_name = record.display_name() or variant_display
        key = f"{display_name.lower()}::{source}"
        if key in seen:
            continue
        seen.add(key)
        ordered.append(
            RosterCandidate(
                name=display_name,
                score=score,
                confidence=_score_to_confidence(score),
                source=source,
            )
        )
        if len(ordered) >= limit:
            break
    return tuple(ordered)


def _score_to_confidence(score: float) -> str:
    if score >= 0.9:
        return "high"
    if score >= 0.8:
        return "medium"
    if score >= 0.7:
        return "low"
    return "none"


def _score_match(
    value_norm: str,
    value_tokens: Tuple[str, ...],
    variant_norm: str,
    variant_tokens: Tuple[str, ...],
) -> float:
    if not value_norm or not variant_norm:
        return 0.0
    sequence_ratio = SequenceMatcher(None, value_norm, variant_norm).ratio()
    if not value_tokens or not variant_tokens:
        return sequence_ratio
    intersection = len(set(value_tokens) & set(variant_tokens))
    denominator = max(len(value_tokens), len(variant_tokens))
    token_score = intersection / denominator if denominator else 0.0
    return 0.4 * sequence_ratio + 0.6 * token_score


def _enumerate_record_variants(record: StudentRecord) -> Iterable[Tuple[str, Tuple[str, ...], str]]:
    names = [
        record.full_name,
        f"{record.first_name} {record.last_name}",
        f"{record.last_name} {record.first_name}",
        record.first_name,
        record.last_name,
    ]
    seen: set[str] = set()
    for name in names:
        normalized = _normalize_for_match(name)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        tokens = tuple(normalized.split())
        yield normalized, tokens, name.strip()


def _expand_handwritten_candidates(
    inputs: Sequence[Tuple[str, str]]
) -> Iterable[Tuple[str, Tuple[str, ...], str, str]]:
    for raw_value, source in inputs:
        if not raw_value:
            continue
        for variant in _expand_single_handwritten(raw_value):
            normalized = _normalize_for_match(variant)
            if not normalized:
                continue
            tokens = tuple(normalized.split())
            yield normalized, tokens, variant, source


def _expand_single_handwritten(raw_value: str) -> List[str]:
    raw_value = raw_value.strip()
    if not raw_value:
        return []

    parts: List[str] = []
    collapsed = _collapse_spaces(raw_value)
    parts.append(collapsed)

    split_parts = [
        _strip_label(part)
        for part in re.split(r"[\n,;/]+", raw_value)
        if _strip_label(part)
    ]
    if split_parts:
        if len(split_parts) >= 2:
            parts.append(_collapse_spaces(" ".join(split_parts)))
            parts.append(_collapse_spaces(" ".join(reversed(split_parts))))
        parts.extend(split_parts)

    tokens = collapsed.split()
    if len(tokens) == 2:
        swapped = f"{tokens[1]} {tokens[0]}"
        parts.append(swapped)

    deduped: List[str] = []
    seen: set[str] = set()
    for value in parts:
        key = value.lower()
        if key not in seen:
            deduped.append(value)
            seen.add(key)
    return deduped


def _strip_label(value: str) -> str:
    value = value.strip()
    value = re.sub(
        r"^\s*(nom|pr[eé]nom|name|first\s*name|last\s*name|surname)\s*[:\-]?\s*",
        "",
        value,
        flags=re.IGNORECASE,
    )
    return _collapse_spaces(value)


def _normalize_for_match(value: str | None) -> str:
    if not value:
        return ""
    normalized = unicodedata.normalize("NFKD", value)
    without_marks = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    cleaned = re.sub(r"[’'`]", " ", without_marks)
    cleaned = re.sub(r"[^0-9A-Za-z\s\-]", " ", cleaned)
    cleaned = cleaned.replace("-", " ")
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip().lower()


def _collapse_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def _row_to_student(row: Dict[str, str]) -> Optional[StudentRecord]:
    if not row:
        return None

    def _lookup(keys: Iterable[str]) -> str:
        for column, value in row.items():
            if not column:
                continue
            column_key = _normalize_for_match(column)
            if column_key in keys:
                return str(value or "").strip()
        return ""

    last_name = _lookup(_LAST_NAME_KEYS)
    first_name = _lookup(_FIRST_NAME_KEYS)
    full_name = _lookup(_FULL_NAME_KEYS)
    if not full_name and first_name and last_name:
        full_name = f"{first_name} {last_name}".strip()
    if not first_name and not last_name and full_name:
        tokens = full_name.split()
        if len(tokens) >= 2:
            first_name = tokens[0]
            last_name = " ".join(tokens[1:])
    email = _lookup(_EMAIL_KEYS) or None

    if not any([first_name, last_name, full_name]):
        return None

    return StudentRecord(
        first_name=_collapse_spaces(first_name),
        last_name=_collapse_spaces(last_name),
        full_name=_collapse_spaces(full_name),
        email=email.strip() if isinstance(email, str) else email,
        raw=row,
    )


def _read_csv_rows(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(2048)
        handle.seek(0)
        delimiter = _detect_delimiter(sample)
        reader = csv.DictReader(handle, delimiter=delimiter)
        return [_clean_row(row) for row in reader if any(_value.strip() for _value in row.values())]


def _detect_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,")
        return dialect.delimiter
    except csv.Error:
        return ";" if sample.count(";") > sample.count(",") else ","


def _read_xlsx_rows(path: Path) -> List[Dict[str, str]]:
    if load_workbook is None:
        raise RosterError(
            "Reading XLSX rosters requires the optional 'openpyxl' package. "
            "Install it with 'pip install openpyxl' and retry."
        )
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        headers: List[str] = []
        rows: List[Dict[str, str]] = []
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            values = [("" if cell is None else str(cell)).strip() for cell in row]
            if index == 0:
                headers = values
                continue
            if not headers:
                continue
            row_dict = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
            if any(value for value in row_dict.values()):
                rows.append(_clean_row(row_dict))
        return rows
    finally:
        workbook.close()


def _clean_row(row: Dict[str, str]) -> Dict[str, str]:
    return {key: (value or "").strip() for key, value in row.items() if key}
